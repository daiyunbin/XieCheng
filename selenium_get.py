# -*- coding:utf-8 -*-

from selenium import webdriver
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from lxml import etree
import random
import os
import time
import datetime
import re
import io, sys


# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')


def create_file():
    if os.path.exists('携程.xlsx'):
        pass
    else:
        wb = Workbook()
        ws1 = wb.create_sheet('Mysheet')
        ws1.append(['ID', '酒店名', '电话', '房型', '市', '区', '路', '周边区域'])
        ws2 = wb.create_sheet('Mysheet1')
        ws2.append(['ID', '酒店名', '日期', '房型', '床型', '早餐', '在线付价格', '到店付价格', '房量'])
        # ws = wb.active
        # ws.append(['ID', '酒店名', '电话', '时间', '房型', '床型', '早餐', '到店付价格', '在线付价格', '房量', '市', '区', '路', '关键词'])
        wb.save('携程.xlsx')


def into_sheet1(info):
    wb = load_workbook('携程.xlsx')
    ws1 = wb.get_sheet_by_name('Mysheet')
    ws1.append(info)
    wb.save('携程.xlsx')


def into_sheet2(info):
    wb = load_workbook('携程.xlsx')
    ws1 = wb.get_sheet_by_name('Mysheet1')
    ws1.append(info)
    wb.save('携程.xlsx')


# 随机的user-agent
def ua():
    l = [
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
        'Opera/9.80 (Windows NT 6.1; U; zh-cn) Presto/2.9.168 Version/11.50',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 2.0.50727; SLCC2; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.3; .NET4.0C; Tablet PC 2.0; .NET4.0E)',
        'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1)',
        'Opera/9.80 (X11; Linux i686; U; ru) Presto/2.8.131 Version/11.11',
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.67 Safari/537.36",
        "Mozilla/5.0 (X11; OpenBSD i386) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.125 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1944.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2309.372 Safari/537.36",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2117.157 Safari/537.36",
    ]
    ua = random.choice(l)
    return ua


def get_url(name):
    """根据酒店名，获取酒店url"""
    url = 'https://m.ctrip.com/restapi/h5api/globalsearch/search?action=online&source=globalonline&keyword={}'.format(
        name)
    headers = {'User-Agent': ua()}
    rep = requests.get(url, headers=headers)
    # print(rep.json())
    hotel_url = rep.json()['data'][0]['url']
    return hotel_url


# 数据写入表一,'ID', '酒店名', '电话', '房型', '市', '区', '路', '周边区域'
# id + 房型为主键，用来去重
def to_sql_sheet1(info):
    connection = pymysql.connect(host='47.105.63.140', port=3306, user='root', password='root', db='携程')
    cursor = connection.cursor()
    # 判断表是否存在
    cursor.execute('show tables;')
    tables = [cursor.fetchall()]
    table_list = re.findall('(\'.*?\')', str(tables))
    table_list = [re.sub("'", '', each) for each in table_list]
    if 'sheet1' in table_list:
        pass
    else:
        yuanneiqiye = '''
                    CREATE TABLE IF NOT EXISTS sheet1(
                    idfangxing varchar(255) not null primary key ,
                    hotel_id varchar(255),
                    hotel_name varchar(255),
                    hotel_tel varchar(255),
                    hotel_room varchar(255),
                    shi varchar(255),
                    qu varchar(255),
                    lu varchar(255),
                    zhoubian varchar(255)
                    )DEFAULT CHARSET=utf8;
                    '''
        cursor.execute(yuanneiqiye)
    insert = '''INSERT INTO sheet1(idfangxing, hotel_id, hotel_name, hotel_tel, hotel_room, shi, qu, lu, zhoubian)
                values (%s,%s,%s,%s,%s,%s,%s,%s,%s)on duplicate key update idfangxing=values(idfangxing)'''
    # 从文件中读出数据
    cursor.execute(insert, (info[0], info[1], info[2], info[3], info[4], info[5], info[6], info[7], info[8]))
    connection.commit()
    cursor.close()


# 获取页面
def get_page(user, password, keys):
    create_file()
    option = webdriver.ChromeOptions()
    option.add_experimental_option('excludeSwitches', ['enable-automation'])
    # option.add_argument('headless')
    option.add_argument('User-Agent=' + ua())
    chromedriver = "C:/Program Files (x86)/Google/Chrome/Application/chromedriver.exe"
    os.environ["webdriver.chrome.driver"] = chromedriver
    driver = webdriver.Chrome(chromedriver, options=option)
    driver.get('https://passport.ctrip.com/user/login')
    driver.delete_all_cookies()
    driver.refresh()
    time.sleep(1)
    driver.find_element_by_id('nloginname').send_keys(user)
    time.sleep(1)
    driver.find_element_by_id('npwd').send_keys(password)
    time.sleep(1)
    driver.find_element_by_id('nsubmit').click()
    time.sleep(5)
    driver.find_element_by_id('_allSearchKeyword').send_keys(keys)
    time.sleep(1)
    driver.find_element_by_id('search_button_global').click()
    time.sleep(5)
    handlers = driver.window_handles
    driver.switch_to.window(handlers[-1])
    time.sleep(5)
    page = driver.page_source
    parse_page1(page)
    # 时间日期，当天和第二天
    today = datetime.date.today()
    tomorrow = today + datetime.timedelta(days=1)
    parse_page(page, today)
    for i in range(30):
        try:
            today = tomorrow
            tomorrow = today + datetime.timedelta(days=1)
            # 写入入住时间
            driver.find_element_by_id('cc_txtCheckIn').clear()
            driver.find_element_by_id('cc_txtCheckIn').send_keys(str(today))
            time.sleep(1.5)
            # 写入退房时间
            driver.find_element_by_id('cc_txtCheckOut').clear()
            driver.find_element_by_id('cc_txtCheckOut').send_keys(str(tomorrow))
            time.sleep(1.5)
            driver.find_element_by_id('changeBtn').click()
            time.sleep(3)
            page = driver.page_source
            # print(page)
            parse_page(page, today)
        except Exception as e:
            pass
    driver.close()


# 写入表1
def parse_page1(page):
    html = etree.HTML(page)
    # 地址
    address = html.xpath('//div[@class="adress"]/span[1]/text()')
    address = ''.join([i.replace('\n', '').replace(' ', '') for i in address])
    address2 = html.xpath('//div[@class="adress"]/span[2]/text()')
    address2 = ''.join([i.replace('\n', '').replace(' ', '') for i in address2])
    address3 = html.xpath('//div[@class="adress"]/span[3]/text()')
    address3 = ''.join([i.replace('\n', '').replace(' ', '') for i in address3])
    address4 = html.xpath('//div[@class="adress"]/a/text()')
    address4 = ''.join([i.replace('\n', '').replace(' ', '') for i in address4])
    # print(''.join(address))
    # 酒店id
    # hotel_id = url.split('/')[-1].replace('.html', '')
    hotel_id = html.xpath('//a[@id="linkViewMap"]/@data-hotelid')
    # 酒店名称
    hotel_name = html.xpath('//h2[@itemprop="name"]/text()')
    # print(hotel_id)
    # print(hotel_name)
    # 电话
    tel = ''.join(html.xpath('//span[@id="J_realContact"]/@data-real'))
    tel = re.findall('电话\d+[-]\d+', tel)
    tel = ''.join(tel).replace('电话', '')
    # print(tel)
    # 房间名
    lst = html.xpath('//tr[@data-disable="0"]')
    roomname = ''
    # wb = load_workbook('携程.xlsx')
    # ws = wb.active
    infos = {}
    for i in range(2, len(lst) + 1):
        s = []
        room_name = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[1]/a[2]/text()'.format(i))).strip()
        if room_name != '':
            roomname = room_name
        # print(room_name)
        # 价格
        # room_price = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[2]/@data-pricedisplay|tr[@data-disable="0"][{}]/td[1]/@data-pricedisplay'.format(i,i))).strip()
        room_price = ''.join(html.xpath('//tr[@data-disable="0"][{}]//td/@data-pricedisplay'.format(i, i))).strip()
        # 床型
        bed = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[@class="col3"]/text()'.format(i)))
        # 早餐
        breakfast = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[contains(@class,"col4")]/text()'.format(i)))
        # breakfast = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[3]/text()'.format(i)))
        # print('break', breakfast)
        breakfast = re.findall('\d', breakfast)
        if ''.join(breakfast) == '':
            breakfast = '0'
        else:
            breakfast = ''.join(breakfast)
        # 付款方式
        payment = ''.join(
            html.xpath('//tr[@data-disable="0"][{}]/td[last()]//span[@class="payment_txt"]/text()'.format(i)))
        if str(payment) == '担保':
            payment = '到店付'
        # print(roomname, room_price, bed, breakfast)
        # 房量
        lastroom = ''.join(
            html.xpath('//tr[@data-disable="0"][{}]/td[last()]//div[@class="hotel_room_last"]/text()'.format(i)))
        info = [''.join(hotel_id), ''.join(hotel_name), tel, str(roomname), str(bed), str(breakfast), str(room_price),
                str(lastroom), address, address2, address3, address4]
        sheet = [''.join(hotel_id) + str(roomname), ''.join(hotel_id), ''.join(hotel_name), tel, str(roomname), address,
                 address2, address3, address4]
        infos.update({str(roomname): sheet})
    for k, v in infos.items():
        print(v)
        # into_sheet1(v)
        v = [str(i) for i in v]
        to_sql_sheet1(v)


# 写入sheet2   'ID', '酒店名', '日期', '房型', '床型', '早餐', '在线付价格', '到店付价格', '房量'
# 日期+酒店名+房间+早餐作为主键
def to_sql_sheet2(info):
    connection = pymysql.connect(host='47.105.63.140', port=3306, user='root', password='root', db='携程')
    cursor = connection.cursor()
    # 判断表是否存在
    cursor.execute('show tables;')
    tables = [cursor.fetchall()]
    table_list = re.findall('(\'.*?\')', str(tables))
    table_list = [re.sub("'", '', each) for each in table_list]
    if 'sheet2' in table_list:
        pass
    else:
        yuanneiqiye = '''
                    CREATE TABLE IF NOT EXISTS sheet2(
                    quchong varchar(255) not null primary key ,
                    hotel_id varchar(255),
                    hotel_name varchar(255),
                    hotel_day varchar(255),
                    hotel_room varchar(255),
                    hotel_bed varchar(255),
                    hotel_breakfast varchar(255),
                    hotel_price1 varchar(255),
                    hotel_price2 varchar(255),
                    hotel_roomnum varchar(255),
                    zuidijia varchar(255)
                    )DEFAULT CHARSET=utf8;
                    '''
        cursor.execute(yuanneiqiye)
    try:
        insert = '''INSERT INTO sheet2(quchong, hotel_id,hotel_name,hotel_day,hotel_room,hotel_bed,hotel_breakfast,hotel_price1,hotel_price2,hotel_roomnum,zuidijia)
                    values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) on duplicate key update quchong=values(quchong)'''
        # 从文件中读出数据
        cursor.execute(insert, (
            info[9], info[0], info[1], info[2], info[3], info[4], info[5], str(int(int(info[6]) * 0.85)),
            str(int(int(info[7]) * 0.85)), info[8], str(int(info[10]) * 0.85)))
        connection.commit()
    except Exception as e:
        pass
    finally:
        cursor.close()


# 解析页面，写入表2
def parse_page(page, today):
    html = etree.HTML(page)
    # 地址
    address = html.xpath('//div[@class="adress"]/span[1]/text()')
    address = ''.join([i.replace('\n', '').replace(' ', '') for i in address])
    address2 = html.xpath('//div[@class="adress"]/span[2]/text()')
    address2 = ''.join([i.replace('\n', '').replace(' ', '') for i in address2])
    address3 = html.xpath('//div[@class="adress"]/span[3]/text()')
    address3 = ''.join([i.replace('\n', '').replace(' ', '') for i in address3])
    address4 = html.xpath('//div[@class="adress"]/a/text()')
    address4 = ''.join([i.replace('\n', '').replace(' ', '') for i in address4])
    # print(''.join(address))
    # 酒店id
    # hotel_id = url.split('/')[-1].replace('.html', '')
    hotel_id = html.xpath('//a[@id="linkViewMap"]/@data-hotelid')
    # 酒店名称
    hotel_name = html.xpath('//h2[@itemprop="name"]/text()')
    # print(hotel_id)
    # print(hotel_name)
    # 电话
    tel = ''.join(html.xpath('//span[@id="J_realContact"]/@data-real'))
    tel = re.findall('电话\d+[-]\d+', tel)
    tel = ''.join(tel).replace('电话', '')
    # print(tel)
    # 房间名
    lst = html.xpath('//tr[@data-disable="0"]')
    roomname = ''

    # infos = []
    infos = {}
    for i in range(2, len(lst) + 1):
        s = []
        room_name = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[1]/a[2]/text()'.format(i))).strip()
        if room_name != '':
            roomname = room_name
        # print(room_name)
        # 是否定完
        room_out = ''.join(
            html.xpath('//tr[@data-disable="0"][{}]//div[@class="btns_base22_main"]/text()'.format(i, i))).strip()
        # 价格
        # room_price = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[2]/@data-pricedisplay|tr[@data-disable="0"][{}]/td[1]/@data-pricedisplay'.format(i,i))).strip()
        room_price = ''.join(html.xpath('//tr[@data-disable="0"][{}]//td/@data-pricedisplay'.format(i, i))).strip()
        # 房间订完的话，房价为空，不写入，
        if room_out == '订完':
            room_price = ''
        # 床型
        bed = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[@class="col3"]/text()'.format(i)))
        # 早餐
        breakfast = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[contains(@class,"col4")]/text()'.format(i)))
        # breakfast = ''.join(html.xpath('//tr[@data-disable="0"][{}]/td[3]/text()'.format(i)))
        # print('break', breakfast)
        breakfast = re.findall('\d', breakfast)
        if ''.join(breakfast) == '':
            breakfast = '0'
        else:
            breakfast = ''.join(breakfast)
        # 付款方式
        payment = ''.join(
            html.xpath('//tr[@data-disable="0"][{}]/td[last()]//span[@class="payment_txt"]/text()'.format(i)))
        if str(payment) == '担保':
            payment = '到店付'
        # print(roomname, room_price, bed, breakfast)
        # 房量
        lastroom = ''.join(
            html.xpath('//tr[@data-disable="0"][{}]/td[last()]//div[@class="hotel_room_last"]/text()'.format(i)))
        # if room_out == '订完':
        #     lastroom = '满房'
        info = [''.join(hotel_id), ''.join(hotel_name), str(today), str(roomname), str(bed), str(breakfast),
                str(room_price), str(lastroom), str(today) + ''.join(hotel_name) + str(roomname) + str(breakfast)]
        #  房间类型+早餐类型+到店付/在线付  作为键去重
        infos.update({str(roomname) + str(breakfast) + str(payment): info})
        # for i in infos:
        #     for k in i.keys():
        #         s.append(k)
        # if str(roomname) + str(breakfast)+str(payment) not in s:
        #     infos.append({str(roomname) + str(breakfast)+str(payment): info})
    print('infos', infos)
    copy_l = {}
    for k, v in infos.items():
        copy_l[k] = v
    # print(copy_l)
    # 用一个列表，放用过的key
    key_lst = []
    for k, v in infos.items():
        if '在线付' in k:
            # 如果没在列表里就放进去，在列表里说名已经用过了
            if k not in key_lst:
                key_lst.append(k)
                try:
                    zaixianfu_price = copy_l[k.replace('在线付', '到店付')][6]
                    key_lst.append(k.replace('在线付', '到店付'))
                    if v[6] == copy_l[k.replace('在线付', '到店付')][6] == '':
                        v[7] = '满房'
                    v.insert(6, zaixianfu_price)
                    v = [str(i) for i in v]
                    # 在线付和到店付两个价格的低价
                    if v[6] >= v[7]:
                        if v[7] != '':
                            v.insert(15, v[7])
                        else:
                            v.insert(15, v[6])
                    else:
                        if v[6] != '':
                            v.insert(15, v[6])
                        else:
                            v.insert(15, v[7])
                    to_sql_sheet2(v)
                except Exception as e:
                    if v[6] == '':
                        v[7] = '满房'
                    v.insert(6, '')
                    v = [str(i) for i in v]
                    # 在线付和到店付两个价格的低价
                    if v[6] >= v[7]:
                        if v[7] != '':
                            v.insert(15, v[7])
                        else:
                            v.insert(15, v[6])
                    else:
                        if v[6] != '':
                            v.insert(15, v[6])
                        else:
                            v.insert(15, v[7])
                    to_sql_sheet2(v)
                    # into_sheet2(v)

        if '到店付' in k:
            if k not in key_lst:
                key_lst.append(k)
                try:
                    zaixianfu_price = copy_l[k.replace('到店付', '在线付')][6]
                    key_lst.append(k.replace('到店付', '在线付'))
                    if v[6] == copy_l[k.replace('到店付', '在线付')][6] == '':
                        v[7] = '满房'
                    v.insert(7, zaixianfu_price)
                    v = [str(i) for i in v]
                    # 在线付和到店付两个价格的低价
                    if v[6] >= v[7]:
                        if v[7] != '':
                            v.insert(15, v[7])
                        else:
                            v.insert(15, v[6])
                    else:
                        if v[6] != '':
                            v.insert(15, v[6])
                        else:
                            v.insert(15, v[7])
                    to_sql_sheet2(v)
                    # into_sheet2(v)
                except Exception as e:
                    if v[6] == '':
                        v[7] = '满房'
                    v.insert(7, '')
                    v = [str(i) for i in v]
                    # 在线付和到店付两个价格的低价
                    if v[6] >= v[7]:
                        if v[7] != '':
                            v.insert(15, v[7])
                        else:
                            v.insert(15, v[6])
                            # v.insert(15, v[7])
                    else:
                        if v[6] != '':
                            v.insert(15, v[6])
                        else:
                            v.insert(15, v[7])
                    to_sql_sheet2(v)


def main(keys):
    with open('携程.txt', 'r') as f:
        l = f.readlines()
    username = l[0].replace('\n', '')
    password = l[1].replace('\n', '')
    print(username, password)
    # try:
    # url = get_url(name.replace('\n', ''))
    # get_page(username, password, name)
    create_file()
    option = webdriver.ChromeOptions()
    option.add_experimental_option('excludeSwitches', ['enable-automation'])
    # option.add_argument('headless')
    option.add_argument('User-Agent=' + ua())
    chromedriver = "C:/Program Files (x86)/Google/Chrome/Application/chromedriver.exe"
    os.environ["webdriver.chrome.driver"] = chromedriver
    driver = webdriver.Chrome(chromedriver, options=option)
    driver.minimize_window()
    try:
        driver.get('https://passport.ctrip.com/user/login')
        driver.delete_all_cookies()
        driver.refresh()
        time.sleep(1)
        driver.find_element_by_id('nloginname').send_keys(username)
        time.sleep(1)
        driver.find_element_by_id('npwd').send_keys(password)
        time.sleep(1)
        driver.find_element_by_id('nsubmit').click()
        time.sleep(5)
        driver.find_element_by_id('_allSearchKeyword').send_keys(keys)
        time.sleep(1)
        driver.find_element_by_id('search_button_global').click()
        time.sleep(5)
        handlers = driver.window_handles
        driver.switch_to.window(handlers[-1])
        time.sleep(5)
        page = driver.page_source
        parse_page1(page)
        # 时间日期，当天和第二天
        today = datetime.date.today()
        tomorrow = today + datetime.timedelta(days=1)
        parse_page(page, today)
        for i in range(30):
            try:
                today = tomorrow
                tomorrow = today + datetime.timedelta(days=1)
                # 写入入住时间
                driver.find_element_by_id('cc_txtCheckIn').clear()
                driver.find_element_by_id('cc_txtCheckIn').send_keys(str(today))
                time.sleep(1.5)
                # 写入退房时间
                driver.find_element_by_id('cc_txtCheckOut').clear()
                driver.find_element_by_id('cc_txtCheckOut').send_keys(str(tomorrow))
                time.sleep(1.5)
                driver.find_element_by_id('changeBtn').click()
                time.sleep(3)
                page = driver.page_source
                # print(page)
                parse_page(page, today)
            except Exception as e:
                pass
                # driver.close()
    except Exception:
        pass
    finally:
        driver.close()
    print(keys.replace('\n', ''), '抓取完成')
    # except Exception as e:
    #     print('-*-', e)


def get_hotel_name():
    connection = pymysql.connect(host='47.105.63.140', port=3306, user='root', password='root', db='携程')
    cursor = connection.cursor()
    s = '''select hotel_name from hotel_name'''
    cursor.execute(s)
    data = cursor.fetchall()
    connection.commit()
    cursor.close()
    if len(data) == 0:
        print('no data')
    else:
        connection = pymysql.connect(host='47.105.63.140', port=3306, user='root', password='root', db='携程')
        cursor = connection.cursor()
        s = '''select hotel_name from hotel_name'''
        cursor.execute(s)
        data = cursor.fetchall()
        connection.commit()
        cursor.close()
        for i in range(len(data)):
            print(data[::-1][i][0])
            # main(data[::-1][i][0].replace('"', '').strip())
            keys = data[::-1][i][0].replace('"', '').strip()
            with open('携程.txt', 'r') as f:
                l = f.readlines()
            username = l[0].replace('\n', '')
            password = l[1].replace('\n', '')
            print(username, password)
            # try:
            # url = get_url(name.replace('\n', ''))
            # get_page(username, password, name)
            create_file()
            option = webdriver.ChromeOptions()
            option.add_experimental_option('excludeSwitches', ['enable-automation'])
            # option.add_argument('headless')
            option.add_argument('User-Agent=' + ua())
            chromedriver = "C:/Program Files (x86)/Google/Chrome/Application/chromedriver.exe"
            os.environ["webdriver.chrome.driver"] = chromedriver
            driver = webdriver.Chrome(chromedriver, options=option)
            driver.minimize_window()
            try:
                driver.get('https://passport.ctrip.com/user/login')
                driver.delete_all_cookies()
                driver.refresh()
                time.sleep(1)
                driver.find_element_by_id('nloginname').send_keys(username)
                time.sleep(1)
                driver.find_element_by_id('npwd').send_keys(password)
                time.sleep(1)
                driver.find_element_by_id('nsubmit').click()
                time.sleep(5)
                driver.find_element_by_id('_allSearchKeyword').send_keys(keys)
                time.sleep(1)
                driver.find_element_by_id('search_button_global').click()
                time.sleep(5)
                handlers = driver.window_handles
                driver.switch_to.window(handlers[-1])
                time.sleep(5)
                page = driver.page_source
                parse_page1(page)
                # 时间日期，当天和第二天
                today = datetime.date.today()
                tomorrow = today + datetime.timedelta(days=1)
                parse_page(page, today)
                for i in range(30):
                    try:
                        today = tomorrow
                        tomorrow = today + datetime.timedelta(days=1)
                        # 写入入住时间
                        driver.find_element_by_id('cc_txtCheckIn').clear()
                        driver.find_element_by_id('cc_txtCheckIn').send_keys(str(today))
                        time.sleep(1.5)
                        # 写入退房时间
                        driver.find_element_by_id('cc_txtCheckOut').clear()
                        driver.find_element_by_id('cc_txtCheckOut').send_keys(str(tomorrow))
                        time.sleep(1.5)
                        driver.find_element_by_id('changeBtn').click()
                        time.sleep(3)
                        page = driver.page_source
                        # print(page)
                        parse_page(page, today)
                    except Exception as e:
                        pass
                        # driver.close()
            except Exception:
                pass
            finally:
                driver.close()
            print(keys.replace('\n', ''), '抓取完成')
            # time.sleep(60*60*3)


def new_get():
    with open('new.txt', 'r') as f:
        names = f.readlines()
    for i in names:  # 遍历酒店
        keys = i.replace('\n', '')
        with open('携程.txt', 'r') as f:
            l = f.readlines()
        username = l[0].replace('\n', '')
        password = l[1].replace('\n', '')
        print(username, password)
        # try:
        # url = get_url(name.replace('\n', ''))
        # get_page(username, password, name)
        # create_file()
        option = webdriver.ChromeOptions()
        option.add_experimental_option('excludeSwitches', ['enable-automation'])
        # option.add_argument('headless')
        option.add_argument('User-Agent=' + ua())
        driver = webdriver.Chrome(options=option)
        driver.minimize_window()
        try:
            driver.get('https://passport.ctrip.com/user/login')
            driver.delete_all_cookies()
            driver.refresh()
            time.sleep(1)
            driver.find_element_by_id('nloginname').send_keys(username)
            time.sleep(1)
            driver.find_element_by_id('npwd').send_keys(password)
            time.sleep(1)
            driver.find_element_by_id('nsubmit').click()
            time.sleep(5)
            driver.find_element_by_id('_allSearchKeyword').send_keys(keys)
            time.sleep(1)
            driver.find_element_by_id('search_button_global').click()
            time.sleep(5)
            handlers = driver.window_handles
            driver.switch_to.window(handlers[-1])
            time.sleep(5)
            page = driver.page_source
            print(page)
            parse_page1(page)
            # 时间日期，当天和第二天
            today = datetime.date.today()
            tomorrow = today + datetime.timedelta(days=1)
            parse_page(page, today)
            for i in range(30):
                try:
                    today = tomorrow
                    tomorrow = today + datetime.timedelta(days=1)
                    # 写入入住时间
                    driver.find_element_by_id('cc_txtCheckIn').clear()
                    driver.find_element_by_id('cc_txtCheckIn').send_keys(str(today))
                    time.sleep(1.5)
                    # 写入退房时间
                    driver.find_element_by_id('cc_txtCheckOut').clear()
                    driver.find_element_by_id('cc_txtCheckOut').send_keys(str(tomorrow))
                    time.sleep(1.5)
                    driver.find_element_by_id('changeBtn').click()
                    time.sleep(3)
                    page = driver.page_source
                    # print(page)
                    parse_page(page, today)
                except Exception as e:
                    pass
                    # driver.close()
        except Exception:
            pass
        finally:
            driver.close()
        print(keys.replace('\n', ''), '抓取完成')


if __name__ == '__main__':
    # main('')
    # while True:
    #     try:
    #         get_hotel_name()
    #         time.sleep(30)
    #     except Exception as e:
    #         time.sleep(10)
    #         print('e', e)
    # get_page('18368090197','hxf18368090197','南京威斯汀大酒店')
    new_get()
    # get_hotel_name()
